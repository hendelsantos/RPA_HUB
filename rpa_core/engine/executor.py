from __future__ import annotations

import time
import csv
from collections.abc import Callable
from email.message import EmailMessage
import json
from pathlib import Path
import shutil
import smtplib
import sqlite3
import subprocess
from typing import Any
from urllib.error import HTTPError
from urllib.request import Request, urlopen
import zipfile

from openpyxl import Workbook, load_workbook
from playwright.sync_api import Page, TimeoutError as PlaywrightTimeoutError, sync_playwright

from rpa_core.desktop import DesktopController
from rpa_core.engine.models import Target, WorkflowDefinition
from rpa_core.engine.sandbox import StepSandbox
from rpa_core.variables import normalize_url, render_template


LogFn = Callable[[str, str, dict[str, Any] | None], None]
CancelFn = Callable[[], bool]
SecretResolver = Callable[[str], str | None]
WEB_STEPS = {"goto", "click", "fill", "secret_fill", "select", "press", "wait_for", "assert_text", "download", "screenshot"}
DESKTOP_STEPS = {
    "desktop_move",
    "desktop_click",
    "desktop_double_click",
    "desktop_drag",
    "desktop_type",
    "desktop_press",
    "desktop_hotkey",
    "desktop_screenshot",
    "desktop_wait",
}


class WorkflowExecutionError(RuntimeError):
    def __init__(self, message: str, artifacts: list[Path] | None = None) -> None:
        super().__init__(message)
        self.artifacts = artifacts or []


class WorkflowCancelledError(RuntimeError):
    pass


class WorkflowExecutor:
    def __init__(
        self,
        artifacts_dir: Path,
        headless: bool = False,
        secret_resolver: SecretResolver | None = None,
        sandbox: StepSandbox | None = None,
    ) -> None:
        self.artifacts_dir = artifacts_dir
        self.headless = headless
        self.secret_resolver = secret_resolver
        self.sandbox = sandbox or StepSandbox()
        self._desktop: DesktopController | None = None

    def run(
        self,
        workflow_data: dict[str, Any],
        inputs: dict[str, Any] | None = None,
        log: LogFn | None = None,
        should_cancel: CancelFn | None = None,
    ) -> list[Path]:
        workflow = WorkflowDefinition.model_validate(workflow_data)
        context = {**workflow.inputs, **(inputs or {})}
        artifacts: list[Path] = []
        self.artifacts_dir.mkdir(parents=True, exist_ok=True)

        if any(step.type in WEB_STEPS for step in workflow.steps if not self._is_disabled(step)):
            with sync_playwright() as playwright:
                browser = playwright.chromium.launch(headless=self.headless)
                page = browser.new_page(accept_downloads=True)
                try:
                    artifacts.extend(self._execute_steps(workflow.steps, context, log, page, should_cancel))
                except PlaywrightTimeoutError as exc:
                    screenshot = self.artifacts_dir / "failure.png"
                    page.screenshot(path=screenshot, full_page=True)
                    artifacts.append(screenshot)
                    raise WorkflowExecutionError(f"Timeout executando workflow: {exc}", [screenshot]) from exc
                finally:
                    browser.close()
        else:
            artifacts.extend(self._execute_steps(workflow.steps, context, log, None, should_cancel))

        return artifacts

    def _execute_steps(self, steps, context: dict[str, Any], log: LogFn | None, page: Page | None, should_cancel: CancelFn | None) -> list[Path]:
        artifacts: list[Path] = []
        for index, step in enumerate(steps, start=1):
            if should_cancel and should_cancel():
                self._log(log, "WARN", f"Execucao cancelada antes da etapa {index}.", {"step": index, "status": "CANCELLED"})
                raise WorkflowCancelledError("Execucao cancelada pelo usuario.")
            if self._is_disabled(step):
                self._log(
                    log,
                    "INFO",
                    f"Etapa {index} ignorada: {step.type}",
                    {"step": index, "type": step.type, "status": "SKIPPED", "disabled": True},
                )
                continue
            attempts = step.retry + 1
            for attempt in range(1, attempts + 1):
                started = time.monotonic()
                try:
                    self._log(log, "INFO", f"Executando etapa {index}: {step.type}", {"step": index, "attempt": attempt})
                    created = self._execute_step(page, step, context, index)
                    artifacts.extend(created)
                    self._log(
                        log,
                        "INFO",
                        f"Etapa {index} concluida: {step.type}",
                        {"step": index, "type": step.type, "status": "SUCCESS", "duration_ms": round((time.monotonic() - started) * 1000)},
                    )
                    break
                except Exception as exc:
                    if attempt >= attempts:
                        if page is None:
                            self._log(
                                log,
                                "ERROR",
                                f"Etapa {index} falhou: {step.type} - {exc}",
                                {"step": index, "type": step.type, "status": "FAILED", "duration_ms": round((time.monotonic() - started) * 1000)},
                            )
                            raise
                        friendly = self._friendly_step_error(step, index, exc)
                        evidence = self._capture_failure(page, index)
                        data = {
                            "step": index,
                            "type": step.type,
                            "status": "FAILED",
                            "duration_ms": round((time.monotonic() - started) * 1000),
                            "suggestion": self._step_suggestion(step),
                        }
                        if evidence is not None:
                            data["evidence"] = str(evidence)
                        self._log(
                            log,
                            "ERROR",
                            friendly,
                            data,
                        )
                        raise WorkflowExecutionError(friendly, [evidence] if evidence is not None else []) from exc
                    self._log(log, "WARN", f"Retry da etapa {index}: {exc}", {"step": index, "attempt": attempt})
        return artifacts

    def _execute_step(self, page: Page | None, step, context: dict[str, Any], index: int) -> list[Path]:
        artifacts: list[Path] = []
        self.sandbox.check_timeout(step.timeout_ms)

        if step.type == "goto":
            page = self._require_page(page, step.type)
            if not step.url:
                raise ValueError("Etapa goto requer url.")
            page.goto(normalize_url(render_template(step.url, context)), timeout=step.timeout_ms)

        elif step.type == "click":
            page = self._require_page(page, step.type)
            self._locator(page, step.target).click(timeout=step.timeout_ms)

        elif step.type == "fill":
            page = self._require_page(page, step.type)
            value = render_template(step.value, context) or ""
            self._locator(page, step.target).fill(value, timeout=step.timeout_ms)

        elif step.type == "secret_fill":
            page = self._require_page(page, step.type)
            if not step.secret:
                raise ValueError("Etapa secret_fill requer secret.")
            if self.secret_resolver is None:
                raise ValueError("Executor nao recebeu resolvedor de segredos.")
            value = self.secret_resolver(step.secret)
            if value is None:
                raise ValueError(f"Segredo nao encontrado: {step.secret}")
            self._locator(page, step.target).fill(value, timeout=step.timeout_ms)

        elif step.type == "select":
            page = self._require_page(page, step.type)
            value = render_template(step.value, context) or ""
            self._locator(page, step.target).select_option(value, timeout=step.timeout_ms)

        elif step.type == "press":
            page = self._require_page(page, step.type)
            if not step.key:
                raise ValueError("Etapa press requer key.")
            self._locator(page, step.target).press(step.key, timeout=step.timeout_ms)

        elif step.type == "wait_for":
            page = self._require_page(page, step.type)
            self._locator(page, step.target).wait_for(timeout=step.timeout_ms)

        elif step.type == "assert_text":
            page = self._require_page(page, step.type)
            text = render_template(step.value, context) or ""
            self._locator(page, step.target).filter(has_text=text).wait_for(timeout=step.timeout_ms)

        elif step.type == "download":
            page = self._require_page(page, step.type)
            filename = render_template(step.filename, context) or "download"
            with page.expect_download(timeout=step.timeout_ms) as download_info:
                self._locator(page, step.target).click(timeout=step.timeout_ms)
            download = download_info.value
            path = self.artifacts_dir / filename
            download.save_as(path)
            artifacts.append(path)

        elif step.type == "screenshot":
            page = self._require_page(page, step.type)
            name = render_template(step.name, context) or f"step-{index}"
            path = self.artifacts_dir / f"{name}.png"
            page.screenshot(path=path, full_page=True)
            artifacts.append(path)

        elif step.type == "delay":
            time.sleep(step.timeout_ms / 1000)

        elif step.type == "file_create_folder":
            path = self._path(step.path, context, "Etapa file_create_folder requer path.")
            path.mkdir(parents=True, exist_ok=True)

        elif step.type == "file_copy":
            source = self._path(step.source, context, "Etapa file_copy requer source.")
            destination = self._path(step.destination, context, "Etapa file_copy requer destination.")
            self._ensure_can_write(destination, step.overwrite)
            destination.parent.mkdir(parents=True, exist_ok=True)
            if source.is_dir():
                shutil.copytree(source, destination, dirs_exist_ok=step.overwrite)
            else:
                shutil.copy2(source, destination)

        elif step.type == "file_move":
            source = self._path(step.source, context, "Etapa file_move requer source.")
            destination = self._path(step.destination, context, "Etapa file_move requer destination.")
            self._ensure_can_write(destination, step.overwrite)
            destination.parent.mkdir(parents=True, exist_ok=True)
            shutil.move(str(source), str(destination))

        elif step.type == "file_delete":
            path = self._path(step.path, context, "Etapa file_delete requer path.")
            if path.is_dir():
                shutil.rmtree(path)
            elif path.exists():
                path.unlink()

        elif step.type == "file_write_text":
            path = self._path(step.path, context, "Etapa file_write_text requer path.")
            self._ensure_can_write(path, step.overwrite)
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_text(render_template(step.value, context) or "", encoding="utf-8")

        elif step.type == "file_read_text":
            path = self._path(step.path, context, "Etapa file_read_text requer path.")
            if not step.variable:
                raise ValueError("Etapa file_read_text requer variable.")
            context[step.variable] = path.read_text(encoding="utf-8")

        elif step.type == "file_zip":
            source = self._path(step.source, context, "Etapa file_zip requer source.")
            destination = self._path(step.destination, context, "Etapa file_zip requer destination.")
            self._ensure_can_write(destination, step.overwrite)
            destination.parent.mkdir(parents=True, exist_ok=True)
            self._zip(source, destination)
            artifacts.append(destination)

        elif step.type == "file_unzip":
            source = self._path(step.source, context, "Etapa file_unzip requer source.")
            destination = self._path(step.destination, context, "Etapa file_unzip requer destination.")
            destination.mkdir(parents=True, exist_ok=True)
            with zipfile.ZipFile(source) as archive:
                self._check_zip_members(archive, destination)
                archive.extractall(destination)

        elif step.type == "csv_read":
            path = self._path(step.path, context, "Etapa csv_read requer path.")
            if not step.variable:
                raise ValueError("Etapa csv_read requer variable.")
            with path.open(newline="", encoding="utf-8") as csv_file:
                context[step.variable] = list(csv.DictReader(csv_file, delimiter=step.delimiter or ","))

        elif step.type == "csv_write":
            path = self._path(step.path, context, "Etapa csv_write requer path.")
            self._ensure_can_write(path, step.overwrite)
            rows = self._rows(step, context)
            path.parent.mkdir(parents=True, exist_ok=True)
            fieldnames = list(rows[0].keys()) if rows else []
            with path.open("w", newline="", encoding="utf-8") as csv_file:
                writer = csv.DictWriter(csv_file, fieldnames=fieldnames, delimiter=step.delimiter or ",")
                if fieldnames:
                    writer.writeheader()
                    writer.writerows(rows)
            artifacts.append(path)

        elif step.type == "excel_read":
            path = self._path(step.path, context, "Etapa excel_read requer path.")
            if not step.variable:
                raise ValueError("Etapa excel_read requer variable.")
            workbook = load_workbook(path, data_only=True)
            sheet = workbook[step.sheet] if step.sheet in workbook.sheetnames else workbook.active
            values = list(sheet.iter_rows(values_only=True))
            headers = [str(value or "") for value in values[0]] if values else []
            context[step.variable] = [dict(zip(headers, row)) for row in values[1:]]

        elif step.type == "excel_write":
            path = self._path(step.path, context, "Etapa excel_write requer path.")
            self._ensure_can_write(path, step.overwrite)
            rows = self._rows(step, context)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = step.sheet or "Planilha1"
            headers = list(rows[0].keys()) if rows else []
            if headers:
                sheet.append(headers)
                for row in rows:
                    sheet.append([row.get(header) for header in headers])
            path.parent.mkdir(parents=True, exist_ok=True)
            workbook.save(path)
            artifacts.append(path)

        elif step.type == "api_request":
            if not step.url:
                raise ValueError("Etapa api_request requer url.")
            response = self._api_request(step, context)
            if step.variable:
                context[step.variable] = response
            if step.output_name:
                output_path = self.artifacts_dir / f"{render_template(step.output_name, context)}.json"
                output_path.write_text(json.dumps(response, ensure_ascii=False, indent=2), encoding="utf-8")
                artifacts.append(output_path)

        elif step.type == "email_send":
            self._send_email(step, context)

        elif step.type == "db_query":
            path = self._path(step.path, context, "Etapa db_query requer path para SQLite.")
            if not step.query:
                raise ValueError("Etapa db_query requer query.")
            rows = self._query_sqlite(path, render_template(step.query, context))
            if step.variable:
                context[step.variable] = rows
            if step.output_name:
                output_path = self.artifacts_dir / f"{render_template(step.output_name, context)}.csv"
                self._write_csv(output_path, rows)
                artifacts.append(output_path)

        elif step.type == "folder_wait_for_file":
            folder = self._path(step.path, context, "Etapa folder_wait_for_file requer path.")
            pattern = render_template(step.filename, context) or "*"
            found = self._wait_for_file(folder, pattern, step.timeout_ms)
            if step.variable:
                context[step.variable] = str(found)

        elif step.type == "pdf_from_text":
            path = self._path(step.path, context, "Etapa pdf_from_text requer path.")
            self._ensure_can_write(path, step.overwrite)
            path.parent.mkdir(parents=True, exist_ok=True)
            self._write_simple_pdf(path, render_template(step.value, context) or "")
            artifacts.append(path)

        elif step.type == "command_run":
            if not step.command:
                raise ValueError("Etapa command_run requer command.")
            command = [render_template(step.command, context), *[render_template(arg, context) for arg in step.args]]
            self.sandbox.check_command(command[0])
            cwd = self._path(step.cwd, context, "Etapa command_run recebeu cwd vazio.") if step.cwd else None
            env = {key: render_template(value, context) for key, value in step.env.items()}
            result = subprocess.run(
                command,
                cwd=cwd,
                env=self.sandbox.child_env(env),
                capture_output=True,
                check=False,
                text=True,
                timeout=step.timeout_ms / 1000,
            )
            if step.output_name:
                output_path = self.artifacts_dir / f"{render_template(step.output_name, context)}.txt"
                output_path.write_text((result.stdout or "") + (result.stderr or ""), encoding="utf-8")
                artifacts.append(output_path)
            if result.returncode != 0:
                raise RuntimeError(f"Comando falhou com codigo {result.returncode}: {result.stderr or result.stdout}")

        elif step.type == "desktop_move":
            desktop = self._desktop_controller()
            x, y = self._coordinates(step)
            desktop.move(x, y, step.duration_ms)

        elif step.type == "desktop_click":
            desktop = self._desktop_controller()
            desktop.click(x=step.x, y=step.y, button=step.button, clicks=1)

        elif step.type == "desktop_double_click":
            desktop = self._desktop_controller()
            desktop.click(x=step.x, y=step.y, button=step.button, clicks=2)

        elif step.type == "desktop_drag":
            desktop = self._desktop_controller()
            x, y = self._coordinates(step)
            desktop.drag(x, y, step.duration_ms or 300, step.button)

        elif step.type == "desktop_type":
            desktop = self._desktop_controller()
            desktop.type_text(render_template(step.value, context) or "", step.interval_ms)

        elif step.type == "desktop_press":
            desktop = self._desktop_controller()
            if not step.key:
                raise ValueError("Etapa desktop_press requer key.")
            desktop.press(step.key)

        elif step.type == "desktop_hotkey":
            desktop = self._desktop_controller()
            desktop.hotkey([render_template(key, context) for key in step.keys])

        elif step.type == "desktop_screenshot":
            desktop = self._desktop_controller()
            name = render_template(step.name, context) or f"desktop-{index}"
            path = self.artifacts_dir / f"{name}.png"
            desktop.screenshot(path)
            artifacts.append(path)

        elif step.type == "desktop_wait":
            desktop = self._desktop_controller()
            desktop.wait(step.timeout_ms)

        return artifacts

    def _rows(self, step, context: dict[str, Any]) -> list[dict[str, Any]]:
        if step.variable and isinstance(context.get(step.variable), list):
            return [dict(row) for row in context[step.variable]]
        if step.rows:
            return [dict(row) for row in step.rows]
        if step.value:
            rendered = render_template(step.value, context)
            data = json.loads(rendered)
            if not isinstance(data, list):
                raise ValueError("Dados de tabela devem ser uma lista de objetos.")
            return [dict(row) for row in data]
        return []

    def _api_request(self, step, context: dict[str, Any]) -> dict[str, Any]:
        method = (step.method or "GET").upper()
        body = render_template(step.body, context).encode("utf-8") if step.body else None
        headers = {key: render_template(value, context) for key, value in step.headers.items()}
        request = Request(normalize_url(render_template(step.url, context)), data=body, headers=headers, method=method)
        try:
            with urlopen(request, timeout=step.timeout_ms / 1000) as response:
                raw = response.read().decode("utf-8", errors="replace")
                status = response.status
                response_headers = dict(response.headers.items())
        except HTTPError as exc:
            raw = exc.read().decode("utf-8", errors="replace")
            status = exc.code
            response_headers = dict(exc.headers.items())
        parsed: Any
        try:
            parsed = json.loads(raw)
        except json.JSONDecodeError:
            parsed = raw
        return {"status": status, "headers": response_headers, "body": parsed}

    def _send_email(self, step, context: dict[str, Any]) -> None:
        if not step.host or not step.to or not step.subject:
            raise ValueError("Etapa email_send requer host, to e subject.")
        message = EmailMessage()
        message["From"] = step.username or step.to
        message["To"] = render_template(step.to, context)
        message["Subject"] = render_template(step.subject, context)
        message.set_content(render_template(step.value, context) or "")
        password = self.secret_resolver(step.password_secret) if self.secret_resolver and step.password_secret else None
        with smtplib.SMTP(render_template(step.host, context), step.port, timeout=step.timeout_ms / 1000) as smtp:
            smtp.starttls()
            if step.username and password:
                smtp.login(render_template(step.username, context), password)
            smtp.send_message(message)

    def _query_sqlite(self, path: Path, query: str) -> list[dict[str, Any]]:
        with sqlite3.connect(path) as connection:
            connection.row_factory = sqlite3.Row
            cursor = connection.execute(query)
            if cursor.description is None:
                connection.commit()
                return []
            return [dict(row) for row in cursor.fetchall()]

    def _wait_for_file(self, folder: Path, pattern: str, timeout_ms: int) -> Path:
        deadline = time.monotonic() + (timeout_ms / 1000)
        while time.monotonic() <= deadline:
            matches = sorted(folder.glob(pattern))
            if matches:
                return matches[0]
            time.sleep(0.5)
        raise TimeoutError(f"Arquivo nao encontrado em {folder}: {pattern}")

    def _write_csv(self, path: Path, rows: list[dict[str, Any]]) -> None:
        path.parent.mkdir(parents=True, exist_ok=True)
        fieldnames = list(rows[0].keys()) if rows else []
        with path.open("w", newline="", encoding="utf-8") as csv_file:
            writer = csv.DictWriter(csv_file, fieldnames=fieldnames)
            if fieldnames:
                writer.writeheader()
                writer.writerows(rows)

    def _write_simple_pdf(self, path: Path, text: str) -> None:
        escaped = text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
        lines = escaped.splitlines() or [""]
        content = ["BT", "/F1 12 Tf", "50 790 Td"]
        for index, line in enumerate(lines[:45]):
            if index:
                content.append("0 -16 Td")
            content.append(f"({line[:100]}) Tj")
        content.append("ET")
        stream = "\n".join(content).encode("latin-1", errors="replace")
        objects = [
            b"<< /Type /Catalog /Pages 2 0 R >>",
            b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
            b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>",
            b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
            b"<< /Length " + str(len(stream)).encode("ascii") + b" >>\nstream\n" + stream + b"\nendstream",
        ]
        pdf = bytearray(b"%PDF-1.4\n")
        offsets = []
        for index, obj in enumerate(objects, start=1):
            offsets.append(len(pdf))
            pdf.extend(f"{index} 0 obj\n".encode("ascii"))
            pdf.extend(obj)
            pdf.extend(b"\nendobj\n")
        xref = len(pdf)
        pdf.extend(f"xref\n0 {len(objects) + 1}\n0000000000 65535 f \n".encode("ascii"))
        for offset in offsets:
            pdf.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
        pdf.extend(f"trailer << /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref}\n%%EOF\n".encode("ascii"))
        path.write_bytes(pdf)

    def _require_page(self, page: Page | None, step_type: str) -> Page:
        if page is None:
            raise ValueError(f"Etapa {step_type} requer navegador.")
        return page

    def _path(self, value: str | None, context: dict[str, Any], error: str) -> Path:
        if not value:
            raise ValueError(error)
        path = Path(render_template(value, context)).expanduser()
        return self.sandbox.check_path(path)

    def _check_zip_members(self, archive: zipfile.ZipFile, destination: Path) -> None:
        destination_resolved = destination.resolve()
        for member in archive.namelist():
            target = (destination / member).resolve()
            if target != destination_resolved and destination_resolved not in target.parents:
                raise ValueError(f"Arquivo suspeito dentro do ZIP (fora da pasta de destino): {member}")

    def _ensure_can_write(self, path: Path, overwrite: bool) -> None:
        if path.exists() and not overwrite:
            raise FileExistsError(f"Caminho ja existe: {path}. Marque overwrite=true para substituir.")

    def _desktop_controller(self) -> DesktopController:
        if self._desktop is None:
            self._desktop = DesktopController()
        return self._desktop

    def _coordinates(self, step) -> tuple[int, int]:
        if step.x is None or step.y is None:
            raise ValueError(f"Etapa {step.type} requer x e y.")
        return step.x, step.y

    def _capture_failure(self, page: Page, index: int) -> Path | None:
        path = self.artifacts_dir / f"falha-passo-{index}.png"
        try:
            page.screenshot(path=path, full_page=True)
        except Exception:
            return None
        return path

    def _friendly_step_error(self, step, index: int, exc: Exception) -> str:
        message = str(exc)
        if isinstance(exc, PlaywrightTimeoutError) or "Timeout" in message:
            return f"Passo {index} nao encontrou {self._target_name(step)}. {self._step_suggestion(step)}"
        if "Segredo nao encontrado" in message:
            return f"Passo {index} nao encontrou a credencial '{step.secret}'. Cadastre essa credencial em Credenciais ou troque o nome no passo."
        return f"Passo {index} falhou em {step.type}: {message}"

    def _target_name(self, step) -> str:
        target = step.target
        if target is None:
            return "o item da tela"
        if target.label:
            return f"o campo '{target.label}'"
        if target.role and target.name:
            return f"o botao '{target.name}'"
        if target.text:
            return f"o texto '{target.text}'"
        if target.css:
            return f"o seletor '{target.css}'"
        return "o item da tela"

    def _step_suggestion(self, step) -> str:
        if step.type in {"fill", "secret_fill", "select"}:
            return "Abra este passo e ajuste o nome do campo exatamente como aparece no sistema."
        if step.type in {"click", "download"}:
            return "Abra este passo e ajuste o texto do botao/link, como Baixar, Exportar ou Gerar Excel."
        if step.type == "wait_for":
            return "Troque o texto esperado pelo nome real da aba, menu ou titulo que aparece apos o login."
        if step.type == "goto":
            return "Confira se o endereco do site esta correto."
        return "Confira os dados deste passo e tente novamente."

    def _zip(self, source: Path, destination: Path) -> None:
        with zipfile.ZipFile(destination, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            if source.is_dir():
                for path in source.rglob("*"):
                    if path.is_file():
                        archive.write(path, path.relative_to(source))
            else:
                archive.write(source, source.name)

    def _locator(self, page: Page, target: Target | None):
        if target is None:
            raise ValueError("Etapa requer target.")
        if target.role and target.name:
            return page.get_by_role(target.role, name=target.name)
        if target.label:
            return page.get_by_label(target.label)
        if target.text:
            return page.get_by_text(target.text)
        if target.css:
            return page.locator(target.css)
        raise ValueError("Target precisa ter role+name, label, text ou css.")

    def _log(self, log: LogFn | None, level: str, message: str, data: dict[str, Any] | None = None) -> None:
        if log:
            log(level, message, data)

    @staticmethod
    def _is_disabled(step) -> bool:
        return bool((step.meta or {}).get("disabled"))
